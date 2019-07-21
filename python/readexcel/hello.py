import re
import os
import openpyxl
import time
import datetime
import warnings

warnings.filterwarnings("ignore")


def readTbdtsNo():
    """
    功能：读取Excel，获取行号和单号
    输入：文件路径
    输出：字典{行号：单号}
    其他：
    """
    fileName = "性能beta.xlsx"
    sheetName = "new"
    dic = {}

    logPath = os.curdir + os.sep + fileName
    book = openpyxl.load_workbook(logPath, "r")
    sheet = book.get_sheet_by_name(sheetName)

    for row in range(2, sheet.max_row):
        dic[row] = str(sheet.cell(row, 2).value)

    return dic



def dlFile():
    """
    功能：下载log，重命名，行号_单号_log字段
    输入：字典{行号：单号}，文件路径
    输出：log文件
    其他：
    """
    pass

# 输入毫秒级的时间，转出正常格式的时间
def timeStamp(timeNum):
    timeStamp = float(timeNum/1000)
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    ms = timeNum % 1000
    resTime = otherStyleTime + ":" + str(ms)
    return resTime


def transTimestamp():
    """
    功能：查找log中的时间，并重命名文件
    输入：所有文件："@123456789."
    输出：带时间戳的文件
    其他：
    """
    files = os.listdir(os.curdir)
    print(files)
    for file in files:
        tsStr = re.search(r'(?<=[@]).*?(?=[.])', file)
        if tsStr is not None:
            tsInt = int(tsStr.group(0))
            res = timeStamp(tsInt)
            print(res)



def recordLog():
    """
    功能：输出下载信息
    输入：所有Excel
    输出：已下载信息，下载失败信息，已删除信息。
    其他：
    """
    his = open("history.txt", "a+")
    his.write('his\n')
    his.close()

def crtlFileNum():
    """
    功能：控制文件数量，50个
    输入：
    输出：
    其他：文件格式10_单号
    """
    pass

def readExtraTbdtsNo():
    """
    功能：读取txt文件，下载单号
    输入：
    输出：
    其他：
    """
    pass


"""
功能：
输入：
输出：
其他：
"""

if __name__ == "__main__":
    dic = readTbdtsNo()
    print(dic)
    transTimestamp()
    recordLog()





